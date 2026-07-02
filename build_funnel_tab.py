"""'Funnel (L·C·Spend·CPL)' — per Product×Country×Theme, 4 stacked metric rows across the trend
window. GOOGLE-consistent end-to-end: Leads (Google) / Clicks / Spend / CPL (Google) — ad stats are
Google-only so CPL is honest. Metric-AWARE color: Leads/Clicks up=green; CPL down=green; Spend neutral."""
import os, sys
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from wsm_cfg import W, OUT, CUR_LABEL, TREND_MONTHS, DEADBAND, label, shift, bq_client
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont

bq = bq_client()
leads = {}; clicks = {}; cost = {}
for r in bq.query(f"SELECT product,country,theme,ym,leads_google FROM `{W}.monitor_leads_monthly`").result():
    leads[(r.product, r.country, r.theme, r.ym)] = (r.leads_google or 0)   # Google-only (matches spend)
for r in bq.query(f"SELECT product,country,theme,ym,clicks,cost FROM `{W}.monitor_monthly_raw`").result():
    clicks[(r.product, r.country, r.theme, r.ym)] = (r.clicks or 0); cost[(r.product, r.country, r.theme, r.ym)] = (r.cost or 0)

months = TREND_MONTHS
def py(ym): return shift(ym, -12)

NORM = InlineFont(rFont='Calibri', sz=11); SMALL = InlineFont(rFont='Calibri', sz=8, color='808080')
GREEN = PatternFill('solid', fgColor='C6EFCE'); RED = PatternFill('solid', fgColor='FFC7CE')
HDR = PatternFill('solid', fgColor='305496'); HF = Font(name='Calibri', bold=True, color='FFFFFF')
NUMF = Font(name='Calibri', size=11)
TOPB = Border(top=Side(style='medium', color='BFBFBF'))
def inr(v):
    if v >= 1e7: return f'₹{v/1e7:.1f}Cr'
    if v >= 1e5: return f'₹{v/1e5:.1f}L'
    if v >= 1e3: return f'₹{v/1e3:.0f}k'
    return f'₹{v:.0f}'
def disp(metric, v): return f'{v:,.0f}' if metric in ('Leads', 'Clicks') else inr(v)

wb = load_workbook(OUT, rich_text=True); src = wb[CUR_LABEL]
order = []; contrib = {}; seen = set()
for r in range(2, src.max_row+1):
    k = (src.cell(r, 1).value, src.cell(r, 2).value, src.cell(r, 3).value)
    if k[0] is None or k in seen: continue
    seen.add(k); order.append(k); contrib[k] = src.cell(r, 4).value

if 'Funnel (L·C·Spend·CPL)' in wb.sheetnames: del wb['Funnel (L·C·Spend·CPL)']
ws = wb.create_sheet('Funnel (L·C·Spend·CPL)')
for ci, h in enumerate(['Product', 'Country', 'Theme', 'Contribution', 'Metric'], 1):
    c = ws.cell(1, ci, h); c.fill = HDR; c.font = HF
MCOL0 = 6
for j, ym in enumerate(months):
    c = ws.cell(1, MCOL0+j, label(ym)); c.fill = HDR; c.font = HF; c.alignment = Alignment(horizontal='center')
    ws.column_dimensions[get_column_letter(MCOL0+j)].width = 12
for ci, wd in zip(range(1, 6), [11, 15, 26, 11, 14]): ws.column_dimensions[get_column_letter(ci)].width = wd

def fill_for(metric, yoy):
    if metric == 'Spend': return None                      # spend is neutral — never colored
    if yoy is None or abs(yoy) < DEADBAND: return None     # dead-band: small/zero move = no color
    good = (yoy > 0) if metric in ('Leads', 'Clicks') else (yoy < 0)   # CPL: cheaper = good
    return GREEN if good else RED

def cell(ws, r, col, metric, v, yoy):
    if v is None: return
    cc = ws.cell(r, col); d = disp(metric, v)
    if yoy is None:
        if metric in ('Leads', 'Clicks'): cc.value = int(round(v)); cc.number_format = '#,##0'; cc.font = NUMF
        else: cc.value = d; cc.font = NUMF
    else:
        cc.value = CellRichText([TextBlock(NORM, d+' '), TextBlock(SMALL, f'({yoy:+.0f}%)')])
        f = fill_for(metric, yoy)
        if f is not None: cc.fill = f
    cc.alignment = Alignment(horizontal='center')

def yoy(v, p): return ((v-p)/p*100) if (p and p > 0) else None
r = 2; blocks = 0
for k in order:
    if not any((leads.get((*k, m), 0) > 0 or clicks.get((*k, m), 0) > 0) for m in months): continue
    blocks += 1
    for mi, metric in enumerate(['Leads', 'Clicks', 'Spend', 'CPL']):
        mlabel = {'Leads': 'Leads (Google)', 'CPL': 'CPL (Google)'}.get(metric, metric)
        ws.cell(r, 1, k[0]); ws.cell(r, 2, k[1]); ws.cell(r, 3, k[2]); ws.cell(r, 4, contrib.get(k)); ws.cell(r, 5, mlabel)
        if mi == 0:
            for ci in range(1, MCOL0+len(months)): ws.cell(r, ci).border = TOPB
        for j, ym in enumerate(months):
            L = leads.get((*k, ym)); C = clicks.get((*k, ym)); S = cost.get((*k, ym))
            Lp = leads.get((*k, py(ym))); Cp = clicks.get((*k, py(ym))); Sp = cost.get((*k, py(ym)))
            if metric == 'Leads' and L is not None and L > 0: cell(ws, r, MCOL0+j, 'Leads', L, yoy(L, Lp))
            elif metric == 'Clicks' and C is not None and C > 0: cell(ws, r, MCOL0+j, 'Clicks', C, yoy(C, Cp))
            elif metric == 'Spend' and S is not None and S > 0: cell(ws, r, MCOL0+j, 'Spend', S, yoy(S, Sp))
            elif metric == 'CPL':
                cpl = (S/L) if (L and L >= 1 and S) else None
                cplp = (Sp/Lp) if (Lp and Lp >= 2 and Sp) else None
                cy = yoy(cpl, cplp) if (cpl and L >= 2) else None
                if cpl is not None: cell(ws, r, MCOL0+j, 'CPL', cpl, cy)
        r += 1
ws.freeze_panes = 'F2'
wb.save(OUT)
print(f"[ok] 'Funnel (L·C·Spend·CPL)': {blocks} active themes × 4 metrics = {blocks*4} rows, {len(months)} month cols")
