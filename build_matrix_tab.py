"""Country x Product leads-growth heatmaps. TWO tabs:
  - 'Country x Product (Leads YTD)'            : YTD (Jan..target month) vs same months prior year
  - 'Country x Product (Leads <target month>)' : single month YoY (catches recent dips YTD masks)
Each tab = 3 stacked sections (All / Google / Bing), same row/col order. All windows from wsm_cfg."""
import os, sys
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from wsm_cfg import W, OUT, CUR, YOY, YTD_CUR, YTD_PRI, CUR_LABEL, CUR_MON, DEADBAND, bq_client
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
from collections import defaultdict

bq = bq_client()
def newd(): return {e: defaultdict(float) for e in ('all', 'google', 'bing')}
cur_ytd, pri_ytd, cur_may, pri_may = newd(), newd(), newd(), newd()
for r in bq.query(f"SELECT country,product,ym,leads,leads_google,leads_bing FROM `{W}.monitor_leads_monthly`").result():
    k = (r.country, r.product)
    for d, months in [(cur_ytd, YTD_CUR), (pri_ytd, YTD_PRI), (cur_may, [CUR]), (pri_may, [YOY])]:
        if r.ym in months:
            d['all'][k] += (r.leads or 0); d['google'][k] += (r.leads_google or 0); d['bing'][k] += (r.leads_bing or 0)

countries = set(k[0] for k in cur_ytd['all']) | set(k[0] for k in pri_ytd['all'])
products = set(k[1] for k in cur_ytd['all']) | set(k[1] for k in pri_ytd['all'])
def csum(c): return sum(cur_ytd['all'].get((c, p), 0) for p in products)
def psum(p): return sum(cur_ytd['all'].get((c, p), 0) for c in countries)
countries = [c for c in sorted(countries, key=lambda x: -csum(x)) if csum(c) > 0 or sum(pri_ytd['all'].get((c, p), 0) for p in products) > 0]
products = [p for p in sorted(products, key=lambda x: -psum(x)) if psum(p) > 0 or sum(pri_ytd['all'].get((c, p), 0) for c in countries) > 0]

NORM = InlineFont(rFont='Calibri', sz=11); SMALL = InlineFont(rFont='Calibri', sz=8, color='808080')
HDR = PatternFill('solid', fgColor='305496'); HF = Font(name='Calibri', bold=True, color='FFFFFF')
GREEN = PatternFill('solid', fgColor='C6EFCE'); RED = PatternFill('solid', fgColor='FFC7CE'); GREY = PatternFill('solid', fgColor='EDEDED')
TITLEF = PatternFill('solid', fgColor='305496'); CAL = Font(name='Calibri', size=11); CEN = Alignment(horizontal='center')
def make_cell(cv, pv):
    if (cv or 0) < 1 and (pv or 0) < 1: return None, None
    if (pv or 0) < 1 and (cv or 0) >= 1: return CellRichText([TextBlock(NORM, 'NEW '), TextBlock(SMALL, f'({cv:.0f})')]), GREEN
    g = (cv-pv)/pv*100
    return CellRichText([TextBlock(NORM, f'{g:+.0f}% '), TextBlock(SMALL, f'({cv:.0f})')]), (GREEN if g >= DEADBAND else (RED if g <= -DEADBAND else None))

def render(ws, r0, title, cd, pd):
    tc = ws.cell(r0, 1, title); tc.fill = TITLEF; tc.font = Font(name='Calibri', bold=True, color='FFFFFF')
    for ci in range(2, 3+len(products)): ws.cell(r0, ci).fill = TITLEF
    hr = r0+1
    ws.cell(hr, 1, 'Country \\ Product').fill = HDR; ws.cell(hr, 1).font = HF
    for j, p in enumerate(products):
        c = ws.cell(hr, 2+j, p); c.fill = HDR; c.font = HF; c.alignment = CEN
    allc = 2+len(products); c = ws.cell(hr, allc, 'All products'); c.fill = HDR; c.font = HF; c.alignment = CEN
    r = hr+1
    for ctry in countries:
        ws.cell(r, 1, ctry).font = CAL
        for j, p in enumerate(products):
            t, f = make_cell(cd.get((ctry, p), 0), pd.get((ctry, p), 0))
            cc = ws.cell(r, 2+j)
            if t is not None: cc.value = t; cc.alignment = CEN
            if f is not None: cc.fill = f
        t, f = make_cell(sum(cd.get((ctry, p), 0) for p in products), sum(pd.get((ctry, p), 0) for p in products))
        cc = ws.cell(r, allc); cc.font = CAL
        if t is not None: cc.value = t; cc.alignment = CEN
        if f is not None: cc.fill = f
        r += 1
    ws.cell(r, 1, 'All countries').font = Font(name='Calibri', size=11, bold=True); ws.cell(r, 1).fill = GREY
    for j, p in enumerate(products):
        t, f = make_cell(sum(cd.get((c, p), 0) for c in countries), sum(pd.get((c, p), 0) for c in countries))
        cc = ws.cell(r, 2+j); cc.fill = f or GREY; cc.alignment = CEN
        if t is not None: cc.value = t
    t, f = make_cell(sum(cd.values()), sum(pd.values()))
    cc = ws.cell(r, allc); cc.fill = f or GREY; cc.alignment = CEN
    if t is not None: cc.value = t
    return r+2

wb = load_workbook(OUT, rich_text=True)
def build_sheet(name, cur, pri, lab):
    if name in wb.sheetnames: del wb[name]
    ws = wb.create_sheet(name)
    ws.column_dimensions['A'].width = 20
    for j in range(len(products)): ws.column_dimensions[get_column_letter(2+j)].width = 12
    ws.column_dimensions[get_column_letter(2+len(products))].width = 13
    r = 1
    for eng, elab in [('all', 'ALL ENGINES (Google + Bing)'), ('google', 'GOOGLE'), ('bing', 'BING')]:
        r = render(ws, r, f'{elab} — {lab} leads growth (YoY)', cur[eng], pri[eng])
    ws.freeze_panes = 'B1'

# drop any stale matrix tabs (old month names), then build both
for s in [s for s in wb.sheetnames if s.startswith('Country x Product (Leads')]:
    del wb[s]
build_sheet('Country x Product (Leads YTD)', cur_ytd, pri_ytd, f'YTD (Jan-{CUR_MON})')
build_sheet(f'Country x Product (Leads {CUR_LABEL})', cur_may, pri_may, CUR_LABEL)
wb.save(OUT)
print(f"[ok] matrices built: YTD + {CUR_LABEL} ({len(countries)} countries x {len(products)} products, 3 engine sections each)")
