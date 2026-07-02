"""'Leads Trend (YoY)' — per Product×Country×Theme, 3 stacked rows (Google / Bing / All) across
the trend window (most-recent-first from col E). Cell = engine leads + (YoY%) small grey;
green/red with dead-band. Leads = country-dependent presales/sales metric (monitor_leads_monthly)."""
import os, sys
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from wsm_cfg import W, OUT, CUR_LABEL, TREND_MONTHS, DEADBAND, label, shift, bq_client
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont

bq = bq_client()
G = {}; B = {}; T = {}
for r in bq.query(f"SELECT product,country,theme,ym,leads_google,leads_bing,leads FROM `{W}.monitor_leads_monthly`").result():
    k = (r.product, r.country, r.theme, r.ym)
    G[k] = (r.leads_google or 0); B[k] = (r.leads_bing or 0); T[k] = (r.leads or 0)

months = TREND_MONTHS
def py(ym): return shift(ym, -12)

NORM = InlineFont(rFont='Calibri', sz=11); NORMB = InlineFont(rFont='Calibri', sz=11, b=True)
SMALL = InlineFont(rFont='Calibri', sz=8, color='808080')
CAL = Font(name='Calibri', size=11); CALB = Font(name='Calibri', size=11, bold=True)
NUMF = Font(name='Calibri', size=11)
HDR = PatternFill('solid', fgColor='305496'); HF = Font(name='Calibri', bold=True, color='FFFFFF')
GREEN = PatternFill('solid', fgColor='C6EFCE'); RED = PatternFill('solid', fgColor='FFC7CE')
ALLF = PatternFill('solid', fgColor='F2F2F2')
CEN = Alignment(horizontal='center')

wb = load_workbook(OUT, rich_text=True); src = wb[CUR_LABEL]
order = []; seen = set()
for r in range(2, src.max_row+1):
    k = (src.cell(r, 1).value, src.cell(r, 2).value, src.cell(r, 3).value)
    if k[0] is None or k in seen: continue
    seen.add(k); order.append(k)

if 'Leads Trend (YoY)' in wb.sheetnames: del wb['Leads Trend (YoY)']
ws = wb.create_sheet('Leads Trend (YoY)')
for ci, h in enumerate(['Product', 'Country', 'Theme', 'Engine'], 1):
    c = ws.cell(1, ci, h); c.fill = HDR; c.font = HF; c.alignment = (CEN if ci == 4 else Alignment(horizontal='left'))
MCOL0 = 5
for j, ym in enumerate(months):
    c = ws.cell(1, MCOL0+j, label(ym)); c.fill = HDR; c.font = HF; c.alignment = CEN
    ws.column_dimensions[get_column_letter(MCOL0+j)].width = 12
for ci, wd in zip(range(1, 5), [11, 15, 28, 9]): ws.column_dimensions[get_column_letter(ci)].width = wd

def put(r, col, v, yoy, bold=False):
    if v is None or v < 0.5: return
    cc = ws.cell(r, col); n = NORMB if bold else NORM
    if yoy is None:
        cc.value = int(round(v)); cc.font = (CALB if bold else NUMF); cc.number_format = '#,##0'
    else:
        cc.value = CellRichText([TextBlock(n, f"{v:.0f} "), TextBlock(SMALL, f"({yoy:+.0f}%)")])
        if abs(yoy) >= DEADBAND: cc.fill = GREEN if yoy > 0 else RED
    cc.alignment = CEN
def yoy(cur, prior): return ((cur-prior)/prior*100) if (prior and prior >= 0.5) else None

r = 2
for k in order:
    if not any(T.get((*k, m), 0) > 0 for m in months): continue
    for eng, src_d, bold in [('Google', G, False), ('Bing', B, False), ('All', T, True)]:
        ws.cell(r, 1, k[0]).font = (CALB if bold else CAL); ws.cell(r, 2, k[1]).font = (CALB if bold else CAL)
        ws.cell(r, 3, k[2]).font = (CALB if bold else CAL)
        ec = ws.cell(r, 4, eng); ec.font = (CALB if bold else CAL); ec.alignment = CEN
        if bold:
            for ci in range(1, MCOL0+len(months)): ws.cell(r, ci).fill = ALLF
        for j, ym in enumerate(months):
            put(r, MCOL0+j, src_d.get((*k, ym)), yoy(src_d.get((*k, ym), 0), src_d.get((*k, py(ym)), 0)), bold=bold)
        r += 1
ws.freeze_panes = 'E2'
wb.save(OUT)
print(f"[ok] 'Leads Trend (YoY)': {r-2} rows ({(r-2)//3} themes × 3 engines), {len(months)} month cols")
