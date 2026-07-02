"""Add a 'Monitor — Top Movers' tab to the workbook copy: stratified digest (top 3 cells/country
by severity) from Ads_data_WSM.monitor_cells. Leads-primary, so 'everywhere' visibility."""
import google.auth
from google.cloud import bigquery
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
PROJ='it-security-online-marketing'; W=f'{PROJ}.Ads_data_WSM'
OUT='/Users/arun-8846/Downloads/Monitor/Continuous Monitoring AKA Big Boss - Monitor.xlsx'
cred,_=google.auth.default(); cred=cred.with_quota_project(PROJ)
bq=bigquery.Client(project=PROJ, credentials=cred)
rows=list(bq.query(f"""
  SELECT product,country,theme,severity,cell_text,
         ROW_NUMBER() OVER (PARTITION BY country ORDER BY severity DESC) rk,
         MAX(severity) OVER (PARTITION BY country) cmax
  FROM `{W}.monitor_cells` QUALIFY rk<=3
""").result())
rows.sort(key=lambda r:(-r.cmax, r.country, -r.severity))
wb=load_workbook(OUT, rich_text=True)
if 'Monitor — Top Movers' in wb.sheetnames: del wb['Monitor — Top Movers']
ws=wb.create_sheet('Monitor — Top Movers')
hdrs=['Country','Product','Theme','Signals (May vs Apr)']
for ci,h in enumerate(hdrs,1):
    c=ws.cell(1,ci,h); c.font=Font(bold=True,color='FFFFFF'); c.fill=PatternFill('solid',fgColor='305496')
for ci,wd in enumerate([18,12,28,70],1): ws.column_dimensions[chr(64+ci)].width=wd
r=2
for x in rows:
    ws.cell(r,1,x.country); ws.cell(r,2,x.product); ws.cell(r,3,x.theme)
    c=ws.cell(r,4,x.cell_text); c.alignment=Alignment(wrap_text=True,vertical='top')
    r+=1
wb.save(OUT)
print(f"[ok] added 'Monitor — Top Movers' tab: {len(rows)} rows across {len(set(x.country for x in rows))} countries")
print("tabs now:", wb.sheetnames)
