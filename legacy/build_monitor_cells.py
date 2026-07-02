"""
Monitor v1 — build leads-primary signal cells and write them into the local workbook as a new
column (sandbox before Zoho). Headline = LEADS (themes_firstlast_semroi); then CPL, impression
share, CTR, clicks. GAds conversions are DEMOTED (noisy vs real leads). Top-3 lines/cell.

Reads Ads_data_WSM.monitor_monthly + monitor_leads_monthly. Writes:
  - Ads_data_WSM.monitor_cells (BQ)
  - a COPY of the xlsx with a new 'Paid-Search Monitor' column (original untouched)
Usage: python build_monitor_cells.py [--month YYYY-MM]   (default 2026-05, the local tab)
"""
import os, sys, math, argparse, shutil
import google.auth
from google.cloud import bigquery
import pandas as pd
from openpyxl import load_workbook

PROJ='it-security-online-marketing'; W=f'{PROJ}.Ads_data_WSM'
XLSX='/Users/arun-8846/Downloads/Monitor/Continuous Monitoring AKA Big Boss.xlsx'
OUT ='/Users/arun-8846/Downloads/Monitor/Continuous Monitoring AKA Big Boss - Monitor.xlsx'
TAB='May 2026'; COL_HEADER='Paid-Search Monitor'
ap=argparse.ArgumentParser(); ap.add_argument('--month', default='2026-05'); a=ap.parse_args()
CUR=a.month
cred,_=google.auth.default(); cred=cred.with_quota_project(PROJ)
bq=bigquery.Client(project=PROJ, credentials=cred)

SRCTBL='monitor_monthly_raw'   # option-2 self-sufficient source (has current month; no dashboard dependency)
months=[r.ym for r in bq.query(f"SELECT DISTINCT ym FROM `{W}.{SRCTBL}` ORDER BY ym").result()]
PRI=months[months.index(CUR)-1]
print(f"[plan] {CUR} vs {PRI}")

def load(tbl):
    d={}
    for r in bq.query(f"SELECT * FROM `{W}.{tbl}` WHERE ym IN ('{CUR}','{PRI}')").result():
        d.setdefault((r.product,r.country,r.theme),{})[r.ym]=r
    return d
mon=load(SRCTBL); lds=load('monitor_leads_monthly')
keys=set(mon)|set(lds)

def pct(c,p):
    if p in (None,0) or c is None: return None
    return (c-p)/p*100.0
def arrow(x): return '🔺' if x>0 else '🔻'
def fpct(p):
    if p is None: return ''
    if abs(p)>=1000: return f"{'+' if p>0 else '-'}{abs(p)/100:.0f}×"
    return f"{p:+.0f}%"

# thresholds (MEDIUM in-cell)
G=dict(leadpct=25, leadabs=5, leadbase=8, cplpct=25, isp=5.0, ctr=1.5, clk=25, volpct=30, imp=300, cost=20000, imprfloor=500)

def build(mc,mp,lc,lp):
    out=[]
    # ---- LEADS (headline) ----
    lc_v=(lc.leads if lc else None); lp_v=(lp.leads if lp else None)
    if lc_v is not None and lp_v is not None:
        dl=pct(lc_v,lp_v); dla=(lc_v or 0)-(lp_v or 0)
        if (lp_v or 0) < 3 and (lc_v or 0) >= G['leadbase']:
            out.append((100*(lc_v or 0)+1000, f"🆕 leads scaled {lp_v:.0f}→{lc_v:.0f}"))
        elif dl is not None and abs(dl)>=G['leadpct'] and abs(dla)>=G['leadabs'] and max(lc_v or 0,lp_v or 0)>=G['leadbase']:
            out.append((100*abs(dla), f"{arrow(dla)} leads {fpct(dl)} ({lp_v:.0f}→{lc_v:.0f})"))
    # ---- CPL (cost per lead) — needs a real lead base ----
    if mc and mp and lc and lp and (lc.leads or 0)>=G['leadbase'] and (lp.leads or 0)>=G['leadbase']:
        cpl_c=(mc.cost or 0)/lc.leads if lc.leads else None; cpl_p=(mp.cost or 0)/lp.leads if lp.leads else None
        dcpl=pct(cpl_c,cpl_p)
        if dcpl is not None and abs(dcpl)>=G['cplpct']:
            out.append((60*abs((lc.leads or 0)-(lp.leads or 0))+20, f"{'🔻' if dcpl>0 else '🔺'} CPL {fpct(dcpl)} (cost/lead {'up'if dcpl>0 else'down'})"))
    # ---- ad mechanics (supporting): IS, CTR, clicks ----
    if mc and mp:
        dclk=pct(mc.clicks,mp.clicks); dclk_a=(mc.clicks or 0)-(mp.clicks or 0)
        dsis=(mc.sis-mp.sis)*100 if mc.sis is not None and mp.sis is not None else None
        dslir=(mc.slir-mp.slir)*100 if mc.slir is not None and mp.slir is not None else None
        dctr=(mc.ctr-mp.ctr)*100 if mc.ctr is not None and mp.ctr is not None else None
        if dsis is not None and abs(dsis)>=G['isp'] and max(mc.impr or 0,mp.impr or 0)>=G['imprfloor']:
            lbc=max(0.0,1-(mc.sis or 0)-(mc.slir or 0)); lbp=max(0.0,1-(mp.sis or 0)-(mp.slir or 0)); dbud=(lbc-lbp)*100
            drv=("losing to rank" if (dslir or 0)>=dbud and (dslir or 0)>0 else ("losing to budget" if dbud>0 else "")) if dsis<0 else ("less rank loss" if (dslir or 0)<0 else ("budget freed up" if dbud<0 else ""))
            out.append((abs(dsis)*math.sqrt(max(mc.impr or 1,1))/5, f"{arrow(dsis)} search IS {dsis:+.0f}pp ({mp.sis*100:.0f}%→{mc.sis*100:.0f}%){' — '+drv if drv else ''}"))
        if dctr is not None and abs(dctr)>=G['ctr'] and max(mc.impr or 0,mp.impr or 0)>=G['imprfloor']:
            out.append((abs(dctr)*math.sqrt(max(mc.impr or 1,1))/7, f"{arrow(dctr)} CTR {dctr:+.1f}pp ({mp.ctr*100:.1f}%→{mc.ctr*100:.1f}%)"))
        if dclk is not None and abs(dclk)>=G['volpct'] and abs(dclk_a)>=G['clk']:
            out.append((abs(dclk_a)*0.5, f"{arrow(dclk_a)} clicks {fpct(dclk)} ({int(mp.clicks)}→{int(mc.clicks)})"))
    out.sort(key=lambda x:-x[0])
    return [t for _,t in out[:3]], (out[0][0] if out else 0)

rows=[]
for k in keys:
    m=mon.get(k,{}); l=lds.get(k,{})
    lines,sev=build(m.get(CUR),m.get(PRI),l.get(CUR),l.get(PRI))
    if lines:
        rows.append(dict(product=k[0],country=k[1],theme=k[2],severity=round(sev,1),cell_text="\n".join("- "+x for x in lines)))
df=pd.DataFrame(rows)
bq.load_table_from_dataframe(df, f"{W}.monitor_cells",
    job_config=bigquery.LoadJobConfig(write_disposition="WRITE_TRUNCATE")).result()
print(f"[ok] {len(df)} signal cells -> {W}.monitor_cells")

# ---- write into a COPY of the workbook ----
cells={(r.product,r.country,r.theme):r.cell_text for r in df.itertuples()}
shutil.copyfile(XLSX, OUT)
wb=load_workbook(OUT, rich_text=True); ws=wb[TAB]
hdr={ws.cell(1,ci).value:ci for ci in range(1,ws.max_column+1)}
pi,ci_,ti=hdr['Product'],hdr['Country'],hdr['Theme']
COL=hdr['Dec']+1   # snap to O, right after the Dec dot column (no gap)
ws.cell(1,COL,value=COL_HEADER)
filled=0; matched=0
for r in range(2, ws.max_row+1):
    key=(ws.cell(r,pi).value, ws.cell(r,ci_).value, ws.cell(r,ti).value)
    if key in cells:
        ws.cell(r,COL,value=cells[key]); filled+=1
    if key in cells or key in keys: matched+=1
wb.save(OUT)
from openpyxl.utils import get_column_letter
print(f"[ok] wrote column {get_column_letter(COL)} '{COL_HEADER}' -> {os.path.basename(OUT)}")
print(f"     {filled} rows populated (of {len(df)} signal cells)")

print("\n--- ADAP · United States rows as written ---")
for r in df[(df['product']=='ADAP')&(df['country']=='United States')].sort_values('severity',ascending=False).head(6).itertuples():
    print(f"{r.theme}:")
    for ln in r.cell_text.split("\n"): print("   "+ln)
