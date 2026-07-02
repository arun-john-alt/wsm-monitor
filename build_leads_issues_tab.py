"""'Leads Issues (by DRI)' — monthly leads-accountability report, ARRANGED BY DRI.
DRI -> Country -> Product -> Theme. Issue = theme's CONSOLIDATED leads >= decline_pct below trailing
3-month norm (with floors). E = leads + YoY%; I = vs 3-mo%. Google/Bing cols show the engine split.
Comments = spend/clicks always + CTR/IS when moved + change-history actions. All config from wsm_cfg."""
import os, sys
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from wsm_cfg import (W, G, OUT, CUR, YOY, BASE, CUR_MON, CH_START, CH_END, CH_LABEL,
                     DECLINE, BASE_FLOOR, ABS_FLOOR, DEADBAND, CG, dri, bq_client)
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
from collections import defaultdict

bq = bq_client()
L = {}; GL = {}; BL = {}
for r in bq.query(f"SELECT product,country,theme,ym,leads,leads_google,leads_bing FROM `{W}.monitor_leads_monthly`").result():
    k = (r.product, r.country, r.theme, r.ym); L[k] = (r.leads or 0); GL[k] = (r.leads_google or 0); BL[k] = (r.leads_bing or 0)
def v(d, k, ym): return d.get((*k, ym), 0)
def a3(d, k): return sum(v(d, k, m) for m in BASE) / 3.0
def pctc(cur, base): return ((cur-base)/base*100) if base else None

# --- supporting metrics for Comments (Google-search side; CUR vs 3-mo-avg) ---
_ml = "','".join([CUR] + BASE)
RAW = {}; CHd = {}
for r in bq.query(f"SELECT product,country,theme,ym,clicks,cost,ctr,sis FROM `{W}.monitor_monthly_raw` WHERE ym IN ('{_ml}')").result():
    RAW[(r.product, r.country, r.theme, r.ym)] = r
for r in bq.query(f"SELECT product,country,theme,category,SUM(count) n FROM `{G}.change_events` WHERE date BETWEEN '{CH_START}' AND '{CH_END}' GROUP BY 1,2,3,4").result():
    CHd.setdefault((r.product, r.country, r.theme), {})[r.category] = int(r.n or 0)
def _c3(p, c, t, attr):
    rc = RAW.get((p, c, t, CUR)); cv = getattr(rc, attr) if rc else None
    vals = [getattr(RAW[(p, c, t, m)], attr) for m in BASE if (p, c, t, m) in RAW and getattr(RAW[(p, c, t, m)], attr) is not None]
    return cv, (sum(vals)/len(vals) if vals else None)
def comment(p, c, t):
    parts = []
    for attr, lab in [('cost', 'spend'), ('clicks', 'clicks')]:
        cv, bv = _c3(p, c, t, attr); d = pctc(cv, bv) if (cv is not None and bv) else None
        if d is not None: parts.append(f"{lab} {d:+.0f}%")   # always show spend & clicks
    cv, bv = _c3(p, c, t, 'ctr')
    if cv is not None and bv is not None and abs((cv-bv)*100) >= CG['ctr_pp']: parts.append(f"CTR {(cv-bv)*100:+.1f}pp")
    cv, bv = _c3(p, c, t, 'sis')
    if cv is not None and bv is not None and abs((cv-bv)*100) >= CG['is_pp']: parts.append(f"IS {cv*100:.0f}% ({(cv-bv)*100:+.0f}pp)")
    ch = CHd.get((p, c, t), {}); tail = []
    for cat, lab in [('cpc', 'CPC edits'), ('kw_removed', 'kw removed'), ('kw_added', 'kw added'), ('adtext', 'ad-text edits')]:
        if ch.get(cat): tail.append(f"{ch[cat]} {lab}")
    s = ", ".join(parts)
    if tail: s += (" | " if s else "") + f"{CH_LABEL}: " + ", ".join(tail)
    return s if s else "no material Google-search shifts (check Bing / Display / demand)"

themes_by = defaultdict(list)
for (p, c, t, ym) in list(L.keys()): themes_by[(c, p)].append((p, c, t))
groups = {}
for (c, p), keys in themes_by.items():
    keys = list(set(keys)); issues = []
    for k in keys:
        may = v(L, k, CUR); base = a3(L, k)
        if base >= BASE_FLOOR and may <= base*(1-DECLINE/100) and (base-may) >= ABS_FLOOR:
            issues.append((k, may, base, v(L, k, YOY)))
    if not issues: continue
    issues.sort(key=lambda x: -(x[2]-x[1]))
    groups[(c, p)] = dict(issues=issues, keys=keys, impact=sum(b-m for _, m, b, _ in issues))

dimp = defaultdict(float); cimp = defaultdict(float)
for (c, p), g in groups.items(): dimp[dri(c)] += g['impact']; cimp[c] += g['impact']
order = sorted(groups.keys(), key=lambda cp: (-dimp[dri(cp[0])], dri(cp[0]), -cimp[cp[0]], cp[0], -groups[cp]['impact']))

NORM = InlineFont(rFont='Calibri', sz=11); NORMB = InlineFont(rFont='Calibri', sz=11, b=True)
SMALL = InlineFont(rFont='Calibri', sz=8, color='808080')
CAL = Font(name='Calibri', size=11); CALB = Font(name='Calibri', size=11, bold=True)
HDR = PatternFill('solid', fgColor='305496'); HF = Font(name='Calibri', bold=True, color='FFFFFF')
GREY = PatternFill('solid', fgColor='EDEDED'); GREEN = PatternFill('solid', fgColor='C6EFCE'); RED = PatternFill('solid', fgColor='FFC7CE')
DRIF = PatternFill('solid', fgColor='DDEBF7')
CEN = Alignment(horizontal='center'); LEFT = Alignment(horizontal='left')
def valpct(val, pct, bold=False):
    n = NORMB if bold else NORM
    return CellRichText([TextBlock(n, f"{val} "), TextBlock(SMALL, f"({pct:+.0f}%)")]) if pct is not None else CellRichText([TextBlock(n, f"{val}")])
def smallpct(pct): return CellRichText([TextBlock(SMALL, f"({pct:+.0f}%)")]) if pct is not None else '—'
def engcell(may, base):
    if (may or 0) < 0.5 and (base or 0) < 0.5: return ''
    return valpct(round(may), pctc(may, base))

NC = 10
wb = load_workbook(OUT, rich_text=True)
for nm in ('Monitor — Top Movers', 'Leads Issues (by Country)', 'Leads Issues (by DRI)'):
    if nm in wb.sheetnames: del wb[nm]
ws = wb.create_sheet('Leads Issues (by DRI)')
heads = ['DRI', 'Country', 'Product', 'Theme', f'Leads ({CUR_MON}, YoY)', 'Google', 'Bing', '3-mo avg', 'vs 3-mo', 'Comments']
for ci, h in enumerate(heads, 1):
    c = ws.cell(1, ci, h); c.fill = HDR; c.font = HF; c.alignment = CEN
for ci, wd in zip(range(1, NC+1), [13, 16, 12, 28, 13, 11, 11, 10, 9, 42]): ws.column_dimensions[chr(64+ci)].width = wd

r = 2; last_dri = None
for (c, p) in order:
    g = groups[(c, p)]; d = dri(c); keys = g['keys']
    if d != last_dri:
        for ci in range(1, NC+1): ws.cell(r, ci).fill = DRIF
        hc = ws.cell(r, 1, f"▼ {d}"); hc.font = CALB; r += 1; last_dri = d
    pmay = sum(v(L, k, CUR) for k in keys); pbase = sum(a3(L, k) for k in keys); pyoy = sum(v(L, k, YOY) for k in keys)
    pg = sum(v(GL, k, CUR) for k in keys); pgb = sum(a3(GL, k) for k in keys)
    pb = sum(v(BL, k, CUR) for k in keys); pbb = sum(a3(BL, k) for k in keys)
    pv = pctc(pmay, pbase); pyoyp = pctc(pmay, pyoy)
    for ci in range(1, NC+1): ws.cell(r, ci).fill = GREY; ws.cell(r, ci).font = CALB
    ws.cell(r, 1, d); ws.cell(r, 2, c); ws.cell(r, 3, p); ws.cell(r, 4, '■ ALL THEMES (product)')
    dc = ws.cell(r, 5); dc.value = valpct(round(pmay), pyoyp, bold=True); dc.alignment = CEN   # E = YoY%
    _fc = GREEN if (pyoyp is not None and pyoyp >= DEADBAND) else (RED if (pyoyp is not None and pyoyp <= -DEADBAND) else None)
    if _fc: dc.fill = _fc   # leave GREY (set above) when within dead-band or no prior-year data
    gc = ws.cell(r, 6); gc.value = engcell(pg, pgb); gc.alignment = CEN
    bc = ws.cell(r, 7); bc.value = engcell(pb, pbb); bc.alignment = CEN
    ws.cell(r, 8, round(pbase, 1)).alignment = CEN
    ws.cell(r, 9).value = smallpct(pv); ws.cell(r, 9).alignment = CEN   # I = vs 3-mo
    r += 1
    for k, may, base, yy in g['issues']:
        ws.cell(r, 1, d).font = CAL; ws.cell(r, 2, c).font = CAL; ws.cell(r, 3, p).font = CAL
        tc = ws.cell(r, 4, '   '+k[2]); tc.font = CAL; tc.alignment = LEFT
        vc = ws.cell(r, 5); vc.value = valpct(round(may), pctc(may, yy)); vc.fill = RED; vc.alignment = CEN
        gc = ws.cell(r, 6); gc.value = engcell(v(GL, k, CUR), a3(GL, k)); gc.alignment = CEN; gc.font = CAL
        bc = ws.cell(r, 7); bc.value = engcell(v(BL, k, CUR), a3(BL, k)); bc.alignment = CEN; bc.font = CAL
        ws.cell(r, 8, round(base, 1)).font = CAL; ws.cell(r, 8).alignment = CEN
        ws.cell(r, 9).value = smallpct(pctc(may, base)); ws.cell(r, 9).alignment = CEN
        cm = ws.cell(r, 10, comment(k[0], k[1], k[2])); cm.font = CAL; cm.alignment = Alignment(wrap_text=True, vertical='top')
        r += 1
ws.freeze_panes = 'E2'
wb.save(OUT)
print(f"[ok] 'Leads Issues (by DRI)': {len(groups)} blocks, {sum(len(g['issues']) for g in groups.values())} issues, {len(dimp)} DRIs")
